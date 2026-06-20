"""
📊 Tạo file Excel bảng lương mẫu cho công ty FinTech — NeoFin
Dùng để test gửi vào TaxSentry qua email/Telegram
"""
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

wb = Workbook()

# === COLORS ===
DARK_BLUE = "1F4E79"
MED_BLUE = "2E75B6"
LIGHT_BLUE = "D6E4F0"
ACCENT_GREEN = "548235"
ACCENT_RED = "C00000"
ACCENT_GOLD = "BF8F00"
LIGHT_GRAY = "F2F2F2"
BORDER_GRAY = "B4C6E7"

hdr_font = Font(color="FFFFFF", bold=True, size=11, name="Calibri")
hdr_fill = PatternFill("solid", fgColor=DARK_BLUE)
gray_fill = PatternFill("solid", fgColor=LIGHT_GRAY)
green_font = Font(color=ACCENT_GREEN, bold=True, name="Calibri")
red_font = Font(color=ACCENT_RED, bold=True, name="Calibri")
gold_font = Font(color=ACCENT_GOLD, bold=True, name="Calibri")
title_font = Font(bold=True, size=16, color=DARK_BLUE, name="Calibri")
subtitle_font = Font(bold=True, size=12, color=MED_BLUE, name="Calibri")
thin_border = Border(
    left=Side(style="thin", color=BORDER_GRAY),
    right=Side(style="thin", color=BORDER_GRAY),
    top=Side(style="thin", color=BORDER_GRAY),
    bottom=Side(style="thin", color=BORDER_GRAY),
)
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
money_fmt = '#,##0'


def style_row(ws, row, max_col, font=hdr_font, fill=hdr_fill):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = font
        cell.fill = fill
        cell.alignment = center_align
        cell.border = thin_border

def data_cell(ws, r, c, val, fmt=None, font=None):
    cell = ws.cell(row=r, column=c, value=val)
    cell.border = thin_border
    cell.alignment = Alignment(horizontal="center" if c != 2 else "left", vertical="center")
    if r % 2 == 0:
        cell.fill = gray_fill
    if fmt:
        cell.number_format = fmt
    if font:
        cell.font = font
    return cell


# ════════════════════════════════════════
# SHEET 1: BẢNG LƯƠNG THÁNG 06/2026
# ════════════════════════════════════════
ws = wb.active
ws.title = "Bảng lương T06-2026"
ws.sheet_properties.tabColor = DARK_BLUE

widths = [5, 28, 22, 16, 16, 16, 16, 16, 16, 16, 16, 16, 16]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Title
ws.merge_cells("A1:M1")
ws["A1"] = "BẢNG LƯƠNG THÁNG 06/2026"
ws["A1"].font = title_font
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 40

ws.merge_cells("A2:M2")
ws["A2"] = "CÔNG TY CỔ PHẦN CÔNG NGHỆ TÀI CHÍNH NEOFIN — FinTech Lending & Payment Platform"
ws["A2"].font = subtitle_font
ws["A2"].alignment = Alignment(horizontal="center")
ws.row_dimensions[2].height = 25

ws.merge_cells("A3:M3")
ws["A3"] = "Ngày lập: 30/06/2026 • Kỳ lương: 01/06/2026 → 30/06/2026 • Đơn vị: VND"
ws["A3"].font = Font(size=10, color="888888", italic=True, name="Calibri")
ws["A3"].alignment = Alignment(horizontal="center")
ws.row_dimensions[3].height = 20
ws.row_dimensions[4].height = 8

# Header row
headers = [
    "STT", "Họ và tên", "Chức vụ", "Lương cơ bản", "Phụ cấp CN",
    "Thưởng KPI", "Làm thêm", "Tổng thu nhập",
    "BHXH/BHYT/BHTN", "Thuế TNCN", "Khấu trừ khác",
    "Thực lĩnh", "Ghi chú"
]
for c, h in enumerate(headers, 1):
    ws.cell(row=5, column=c, value=h)
style_row(ws, 5, len(headers))
ws.row_dimensions[5].height = 35

# Employees data
employees = [
    # (name, position, base_salary, allowance, kpi_bonus, overtime)
    ("Nguyễn Hoàng Nam",   "CEO — Giám đốc Điều hành",         85000000, 15000000, 20000000,  5000000),
    ("Trần Minh Quân",     "CFO — Giám đốc Tài chính",          65000000, 12000000, 18000000,  4000000),
    ("Lê Thị Hương Giang", "CTO — Giám đốc Công nghệ",          70000000, 12000000, 25000000,  6000000),
    ("Phạm Anh Tú",        "COO — Giám đốc Vận hành",           60000000, 10000000, 15000000,  3500000),
    ("Hoàng Thị Mai",      "Trưởng phòng Kế toán",              35000000,  8000000, 10000000,  2000000),
    ("Vũ Đức Bình",        "Trưởng phòng Phát triển SP",       40000000,  8000000, 15000000,  3000000),
    ("Đặng Thùy Linh",     "Trưởng phòng Nhân sự",             32000000,  7000000,  8000000,  1500000),
    ("Bùi Quốc Anh",       "Dev Lead — Backend",               35000000,  5000000, 12000000,  3000000),
    ("Ngô Thanh Hà",       "Dev Lead — Frontend",              33000000,  5000000, 12000000,  2500000),
    ("Đỗ Minh Hoàng",      "Senior Blockchain Developer",       38000000,  5000000, 18000000,  4000000),
    ("Lý Văn Phúc",        "Senior Data Engineer",             36000000,  5000000, 15000000,  3000000),
    ("Trương Khánh Vân",   "Senior Product Manager",           35000000,  5000000, 12000000,  2000000),
    ("Mai Thanh Tùng",     "Compliance & Risk Officer",         28000000,  5000000,  8000000,  1500000),
    ("Kiều Minh Đức",      "Fullstack Developer",              25000000,  3000000,  8000000,  2000000),
    ("Phan Hồng Nhung",    "UX/UI Designer",                   22000000,  3000000,  6000000,  1500000),
    ("Dương Văn Hải",      "QA/QC Engineer",                  20000000,  3000000,  5000000,  1000000),
    ("Lâm Thị Ngọc",       "HR Executive",                    18000000,  2500000,  4000000,   500000),
    ("Trịnh Bảo Long",     "Marketing Executive",             18000000,  2500000,  5000000,  1000000),
    ("Hà Văn Sơn",         "Customer Support Lead",           17000000,  2500000,  4000000,  1500000),
    ("Tăng Thúy An",       "Finance Staff",                   16000000,  2000000,  3000000,   500000),
]

INSURANCE_CAP = 36800000
PERSONAL_DEDUCTION = 11000000

def calc_tax(taxable):
    """Tính thuế TNCN lũy tiến từng phần."""
    if taxable <= 0:
        return 0
    brackets = [
        (5000000, 0.05, 0),
        (10000000, 0.10, 250000),
        (18000000, 0.15, 750000),
        (32000000, 0.20, 1650000),
        (52000000, 0.25, 3250000),
        (80000000, 0.30, 5850000),
    ]
    for limit, rate, deduct in brackets:
        if taxable <= limit:
            return round(taxable * rate - deduct)
    return round(taxable * 0.35 - 9850000)

row = 6
for idx, (name, pos, base, allow, kpi, ot) in enumerate(employees, 1):
    total_income = base + allow + kpi + ot
    ins_base = min(base, INSURANCE_CAP)
    insurance = round(ins_base * 0.105)  # BHXH 8% + BHYT 1.5% + BHTN 1%
    taxable = total_income - insurance - PERSONAL_DEDUCTION
    tax = calc_tax(taxable)
    net_pay = total_income - insurance - tax

    note = ""
    if total_income > 100000000:
        note = "Quản lý cấp cao"
    elif kpi > 10000000:
        note = "KPI xuất sắc"

    vals = [idx, name, pos, base, allow, kpi, ot, total_income, insurance, tax, 0, net_pay, note]
    for c, v in enumerate(vals, 1):
        f = money_fmt if c >= 4 else None
        ft = None
        if c == 8:
            ft = gold_font
        elif c == 9:
            ft = Font(color=MED_BLUE, bold=True, name="Calibri")
        elif c == 10:
            ft = red_font
        elif c == 12:
            ft = green_font
        data_cell(ws, row, c, v, f, ft)
    row += 1

# Total row
row += 1
sum_row = row
ws.merge_cells(f"A{row}:C{row}")
ws.cell(row=row, column=1, value="TỔNG CỘNG").font = Font(bold=True, size=12, color=DARK_BLUE, name="Calibri")
ws.cell(row=row, column=1).alignment = Alignment(horizontal="right", vertical="center")
ws.cell(row=row, column=1).border = thin_border
ws.cell(row=row, column=2).border = thin_border
ws.cell(row=row, column=3).border = thin_border

for c in range(4, 14):
    cl = get_column_letter(c)
    ws.cell(row=row, column=c,
            value=f"=SUM({cl}6:{cl}{row-1})")
    ws.cell(row=row, column=c).font = Font(bold=True, size=11, color=DARK_BLUE, name="Calibri")
    ws.cell(row=row, column=c).border = Border(
        left=Side(style="thin", color=BORDER_GRAY),
        right=Side(style="thin", color=BORDER_GRAY),
        top=Side(style="double", color=DARK_BLUE),
        bottom=Side(style="double", color=DARK_BLUE),
    )
    ws.cell(row=row, column=c).number_format = money_fmt
    ws.cell(row=row, column=c).alignment = center_align
    ws.cell(row=row, column=c).fill = PatternFill("solid", fgColor=LIGHT_BLUE)
ws.row_dimensions[row].height = 32

# Signatures
row += 3
sig_items = [
    ("NGƯỜI LẬP BẢNG", 1, 4),
    ("KẾ TOÁN TRƯỞNG", 5, 8),
    ("GIÁM ĐỐC", 9, 13),
]
for label, c_start, c_end in sig_items:
    ws.merge_cells(start_row=row, start_column=c_start, end_row=row, end_column=c_end)
    cell = ws.cell(row=row, column=c_start, value=label)
    cell.font = Font(size=10, color="666666", bold=True, name="Calibri")
    cell.alignment = center_align
    cell.border = Border(bottom=Side(style="medium", color="999999"))

row += 1
for _, c_start, c_end in sig_items:
    ws.merge_cells(start_row=row, start_column=c_start, end_row=row, end_column=c_end)
    cell = ws.cell(row=row, column=c_start, value="(Ký, ghi rõ họ tên)")
    cell.font = Font(size=9, color="999999", italic=True, name="Calibri")
    cell.alignment = center_align


# ════════════════════════════════════════
# SHEET 2: TỔNG HỢP THUẾ & BẢO HIỂM
# ════════════════════════════════════════
ws2 = wb.create_sheet("Tổng hợp thuế-BH")
ws2.sheet_properties.tabColor = ACCENT_GREEN
for i, w in enumerate([5, 35, 18, 20, 20, 20, 20], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

ws2.merge_cells("A1:G1")
ws2["A1"] = "TỔNG HỢP THUẾ TNCN & BẢO HIỂM XÃ HỘI — Tháng 06/2026"
ws2["A1"].font = title_font
ws2["A1"].alignment = Alignment(horizontal="center")
ws2.row_dimensions[1].height = 40

th = ["STT", "Chỉ tiêu", "Số lượng", "Tổng tiền lương", "BHXH (8%)", "BHYT (1.5%)", "BHTN (1%)"]
for c, h in enumerate(th, 1):
    ws2.cell(row=3, column=c, value=h)
style_row(ws2, 3, len(th))

summary = [
    ("Tiền lương đóng BH (≤ 36.8tr/người)",  20, 520000000, 41600000, 7800000, 5200000),
    ("Tiền lương trên 36.8tr (phần chênh)",   20, 180000000, 0, 0, 0),
    ("Tổng quỹ lương tháng 06/2026",          20, 700000000, 41600000, 7800000, 5200000),
]
for idx, (item, count, total, bhxh, bhyt, bhtn) in enumerate(summary, 1):
    r = 4 + idx - 1
    for c, v in enumerate([idx, item, count, total, bhxh, bhyt, bhtn], 1):
        f = money_fmt if c >= 4 else None
        data_cell(ws2, r, c, v, f)

# Total row
r = 8
for c in range(1, 8):
    cl = get_column_letter(c)
    if c == 1:
        data_cell(ws2, r, c, "")
    elif c == 2:
        data_cell(ws2, r, c, "CỘNG")
    elif c == 3:
        data_cell(ws2, r, c, f"=SUM(D4:D6)")
    else:
        data_cell(ws2, r, c, f"=SUM({cl}4:{cl}6)", money_fmt)
    ws2.cell(row=r, column=c).font = Font(bold=True, size=11, color=DARK_BLUE)
    ws2.cell(row=r, column=c).fill = PatternFill("solid", fgColor=LIGHT_BLUE)

# ════════════════════════════════════════
# SHEET 3: THÔNG TIN NHÂN VIÊN
# ════════════════════════════════════════
ws3 = wb.create_sheet("Thông tin nhân viên")
ws3.sheet_properties.tabColor = ACCENT_GOLD
for i, w in enumerate([5, 24, 12, 20, 22, 16, 22, 18], 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

ws3.merge_cells("A1:H1")
ws3["A1"] = "DANH SÁCH NHÂN VIÊN — CÔNG TY CỔ PHẦN CÔNG NGHỆ TÀI CHÍNH NEOFIN"
ws3["A1"].font = title_font
ws3["A1"].alignment = Alignment(horizontal="center")
ws3.row_dimensions[1].height = 40

eh = ["STT", "Họ và tên", "Mã NV", "Phòng ban", "Chức vụ", "Ngày vào", "Hợp đồng", "Tình trạng"]
for c, h in enumerate(eh, 1):
    ws3.cell(row=3, column=c, value=h)
style_row(ws3, 3, len(eh))

emp_data = [
    ("Nguyễn Hoàng Nam",      "NV001", "Ban Giám đốc",    "CEO",       "01/03/2020", "Không thời hạn", "Đang làm việc"),
    ("Trần Minh Quân",        "NV002", "Ban Giám đốc",    "CFO",       "15/06/2020", "Không thời hạn", "Đang làm việc"),
    ("Lê Thị Hương Giang",    "NV003", "Ban Giám đốc",    "CTO",       "01/09/2021", "Không thời hạn", "Đang làm việc"),
    ("Phạm Anh Tú",           "NV004", "Ban Giám đốc",    "COO",       "01/01/2022", "Không thời hạn", "Đang làm việc"),
    ("Hoàng Thị Mai",         "NV005", "Phòng Kế toán",   "Trưởng phòng", "15/03/2021", "Không thời hạn", "Đang làm việc"),
    ("Vũ Đức Bình",           "NV006", "Phòng Sản phẩm",  "Trưởng phòng", "01/06/2021", "Không thời hạn", "Đang làm việc"),
    ("Đặng Thùy Linh",        "NV007", "Phòng Nhân sự",   "Trưởng phòng", "01/09/2021", "Không thời hạn", "Đang làm việc"),
    ("Bùi Quốc Anh",          "NV008", "Phòng Kỹ thuật",  "Dev Lead",   "01/03/2022", "1 năm",         "Đang làm việc"),
    ("Ngô Thanh Hà",          "NV009", "Phòng Kỹ thuật",  "Dev Lead",   "01/06/2022", "1 năm",         "Đang làm việc"),
    ("Đỗ Minh Hoàng",         "NV010", "Phòng Blockchain", "Senior Dev", "01/09/2021", "Không thời hạn", "Đang làm việc"),
    ("Lý Văn Phúc",           "NV011", "Phòng Dữ liệu",   "Senior Eng", "15/01/2022", "Không thời hạn", "Đang làm việc"),
    ("Trương Khánh Vân",      "NV012", "Phòng Sản phẩm",  "Senior PM",  "01/06/2022", "1 năm",         "Đang làm việc"),
    ("Mai Thanh Tùng",        "NV013", "Phòng Pháp chế",  "Compliance", "01/03/2023", "1 năm",         "Đang làm việc"),
    ("Kiều Minh Đức",         "NV014", "Phòng Kỹ thuật",  "Fullstack",  "01/09/2023", "1 năm",         "Đang làm việc"),
    ("Phan Hồng Nhung",       "NV015", "Phòng Thiết kế",  "UX/UI Lead", "15/03/2023", "1 năm",         "Đang làm việc"),
    ("Dương Văn Hải",         "NV016", "Phòng Kỹ thuật",  "QA/QC",      "01/06/2023", "1 năm",         "Đang làm việc"),
    ("Lâm Thị Ngọc",          "NV017", "Phòng Nhân sự",   "HR Exec",    "01/09/2023", "1 năm",         "Đang làm việc"),
    ("Trịnh Bảo Long",        "NV018", "Phòng Marketing",  "Mkt Exec",   "01/01/2024", "1 năm",         "Đang làm việc"),
    ("Hà Văn Sơn",            "NV019", "Phòng CSKH",      "CS Lead",    "01/03/2024", "1 năm",         "Đang làm việc"),
    ("Tăng Thúy An",          "NV020", "Phòng Kế toán",   "Kế toán",    "01/06/2024", "1 năm",         "Đang làm việc"),
]

for idx, (name, code, dept, pos, date, contract, status) in enumerate(emp_data, 1):
    r = 4 + idx - 1
    vals = [idx, name, code, dept, pos, date, contract, status]
    for c, v in enumerate(vals, 1):
        ft = Font(color=ACCENT_GREEN, name="Calibri") if status == "Đang làm việc" else Font(color=ACCENT_RED, name="Calibri")
        data_cell(ws3, r, c, v, font=ft)

# ════════════════════════════════════════
# SAVE
# ════════════════════════════════════════
output_dir = Path("D:/TaxSentry/downloads")
output_dir.mkdir(exist_ok=True)
output_path = output_dir / "BaoCao_TienLuong_Thang06_2026_NeoFin.xlsx"
wb.save(str(output_path))
print(f"✅ File created: {output_path}")
print(f"   Sheets: {wb.sheetnames}")
print(f"   Employees: {len(emp_data)}")
