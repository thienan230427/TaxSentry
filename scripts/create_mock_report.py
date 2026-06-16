import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# --- KHỞI TẠO WORKBOOK ---
wb = Workbook()

# Định nghĩa màu sắc & font chữ chuẩn banker-grade (excel-author skill)
BLUE_FONT = Font(name="Calibri", size=11, color="0000FF")       # Dành cho ô nhập liệu (Inputs)
BLACK_FONT = Font(name="Calibri", size=11, color="000000")      # Dành cho ô công thức (Formulas)
BOLD_FONT = Font(name="Calibri", size=11, bold=True)
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

# Fills
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")            # Màu xanh đậm cho Header
SUBTOTAL_FILL = PatternFill("solid", fgColor="D9E1F2")          # Màu xanh nhạt cho các ô tổng phụ
INPUT_FILL = PatternFill("solid", fgColor="F2F2F2")             # Màu xám nhạt cho ô nhập liệu

# Borders
thin_line = Side(border_style="thin", color="D3D3D3")
double_line = Side(border_style="double", color="000000")
thick_bottom = Side(border_style="medium", color="1F4E79")

cell_border = Border(left=thin_line, right=thin_line, top=thin_line, bottom=thin_line)
header_border = Border(left=thin_line, right=thin_line, top=thin_line, bottom=thick_bottom)
total_border = Border(top=thin_line, bottom=double_line)

align_left = Alignment(horizontal="left", vertical="center")
align_right = Alignment(horizontal="right", vertical="center")
align_center = Alignment(horizontal="center", vertical="center")

# ==========================================
# TAB 1: ASSUMPTIONS (Các giả định tính toán)
# ==========================================
ws_assumptions = wb.active
ws_assumptions.title = "Assumptions"
ws_assumptions.views.sheetView[0].showGridLines = True

# Header
ws_assumptions["A1"] = "BẢNG GIẢ ĐỊNH TÀI CHÍNH & THUẾ"
ws_assumptions["A1"].font = HEADER_FONT
ws_assumptions["A1"].fill = HEADER_FILL
ws_assumptions.merge_cells("A1:C1")
ws_assumptions["A1"].alignment = align_center

# Table Headers
ws_assumptions["A3"] = "Chỉ số giả định"
ws_assumptions["B3"] = "Giá trị"
ws_assumptions["C3"] = "Ghi chú pháp lý"
for col in ["A", "B", "C"]:
    ws_assumptions[f"{col}3"].font = BOLD_FONT
    ws_assumptions[f"{col}3"].fill = SUBTOTAL_FILL
    ws_assumptions[f"{col}3"].border = cell_border

# Data Inputs (Màu xanh dương - Blue Font)
data_assumptions = [
    ("Thuế suất Thuế TNDN", 0.20, "Theo Luật thuế TNDN hiện hành (20%)"),
    ("Hạn mức chi phí Tiếp khách hợp lệ (% Doanh thu)", 0.15, "Chi phí tiếp khách tối đa được trừ khi tính thuế"),
    ("Mức phạt chậm nộp thuế mỗi ngày", 0.0003, "Theo Luật quản lý thuế (0.03%/ngày)"),
]

for idx, (label, val, note) in enumerate(data_assumptions, start=4):
    ws_assumptions[f"A{idx}"] = label
    ws_assumptions[f"A{idx}"].alignment = align_left
    ws_assumptions[f"A{idx}"].border = cell_border
    
    ws_assumptions[f"B{idx}"] = val
    ws_assumptions[f"B{idx}"].font = BLUE_FONT
    ws_assumptions[f"B{idx}"].number_format = "0.0%" if val < 1 else "#,##0"
    ws_assumptions[f"B{idx}"].alignment = align_right
    ws_assumptions[f"B{idx}"].border = cell_border
    ws_assumptions[f"B{idx}"].fill = INPUT_FILL
    
    ws_assumptions[f"C{idx}"] = note
    ws_assumptions[f"C{idx}"].alignment = align_left
    ws_assumptions[f"C{idx}"].border = cell_border

# ==========================================
# TAB 2: INCOME STATEMENT (Báo cáo kinh doanh)
# ==========================================
ws_is = wb.create_sheet("Income_Statement")
ws_is.views.sheetView[0].showGridLines = True

# Header chính
ws_is["A1"] = "BÁO CÁO KẾT QUẢ HOẠT ĐỘNG KINH DOANH"
ws_is["A1"].font = HEADER_FONT
ws_is["A1"].fill = HEADER_FILL
ws_is.merge_cells("A1:D1")
ws_is["A1"].alignment = align_center

# Column Headers
ws_is["A3"] = "Chỉ tiêu kinh doanh (VND)"
ws_is["B3"] = "Tháng 04/2026 (Thực tế)"
ws_is["C3"] = "Tháng 05/2026 (Nhập liệu)"
ws_is["D3"] = "Ghi chú từ Kế toán trưởng"

for col in ["A", "B", "C", "D"]:
    ws_is[f"{col}3"].font = BOLD_FONT
    ws_is[f"{col}3"].fill = SUBTOTAL_FILL
    ws_is[f"{col}3"].border = cell_border

# --- DANH SÁCH DỮ LIỆU ---
# Dữ liệu bao gồm các ô nhập liệu cứng (BLUE) và các ô tính bằng công thức (BLACK)
# Định dạng: (Tên chỉ tiêu, Giá trị T4, Giá trị/Công thức T5, Loại ô 'input'/'formula'/'subtotal'/'total')
rows_is = [
    ("Doanh thu bán hàng và cung cấp dịch vụ", 500000000, 650000000, "input"),
    ("Giá vốn hàng bán", 200000000, 240000000, "input"),
    ("Lợi nhuận gộp", "=B4-B5", "=C4-C5", "subtotal"),  # Doanh thu - Giá vốn
    ("Chi phí bán hàng (Marketing, vận chuyển)", 50000000, 75000000, "input"),
    ("Chi phí lương nhân viên", 80000000, 85000000, "input"),
    ("Chi phí tiếp khách (Hợp lệ có hóa đơn)", 20000000, 25000000, "input"),
    ("Chi phí tiếp khách (KHÔNG có hóa đơn đỏ)", 5000000, 45000000, "input"), # Chi phí nhạy cảm để AI bắt lỗi thuế!
    ("Chi phí thuê văn phòng và dịch vụ khác", 30000000, 30000000, "input"),
    ("Tổng chi phí quản lý & vận hành", "=SUM(B7:B11)", "=SUM(C7:C11)", "subtotal"),
    ("Lợi nhuận kế toán trước thuế", "=B6-B12", "=C6-C12", "subtotal"),
    ("Thuế TNDN phải nộp", "=B13*Assumptions!$B$4", "=C13*Assumptions!$B$4", "formula"), # Liên kết sang tab Assumptions
    ("Lợi nhuận ròng sau thuế", "=B13-B14", "=C13-C14", "total")
]

for idx, (label, val_t4, val_t5, row_type) in enumerate(rows_is, start=4):
    ws_is[f"A{idx}"] = label
    ws_is[f"A{idx}"].border = cell_border
    
    # Định dạng cột Tháng 4 (Thực tế - Inputs lịch sử)
    ws_is[f"B{idx}"] = val_t4
    ws_is[f"B{idx}"].alignment = align_right
    ws_is[f"B{idx}"].border = cell_border
    if str(val_t4).startswith("="):
        ws_is[f"B{idx}"].font = BLACK_FONT
    else:
        ws_is[f"B{idx}"].font = BLUE_FONT
        ws_is[f"B{idx}"].fill = INPUT_FILL
    ws_is[f"B{idx}"].number_format = "#,##0"

    # Định dạng cột Tháng 5 (Nhập liệu hiện tại)
    ws_is[f"C{idx}"] = val_t5
    ws_is[f"C{idx}"].alignment = align_right
    ws_is[f"C{idx}"].border = cell_border
    if str(val_t5).startswith("="):
        ws_is[f"C{idx}"].font = BLACK_FONT
    else:
        ws_is[f"C{idx}"].font = BLUE_FONT
        ws_is[f"C{idx}"].fill = INPUT_FILL
    ws_is[f"C{idx}"].number_format = "#,##0"

    # Styles đặc biệt cho từng loại dòng
    if row_type == "subtotal":
        ws_is[f"A{idx}"].font = BOLD_FONT
        ws_is[f"B{idx}"].font = BOLD_FONT
        ws_is[f"C{idx}"].font = BOLD_FONT
        ws_is[f"A{idx}"].fill = SUBTOTAL_FILL
        ws_is[f"B{idx}"].fill = SUBTOTAL_FILL
        ws_is[f"C{idx}"].fill = SUBTOTAL_FILL
    elif row_type == "total":
        ws_is[f"A{idx}"].font = BOLD_FONT
        ws_is[f"B{idx}"].font = BOLD_FONT
        ws_is[f"C{idx}"].font = BOLD_FONT
        ws_is[f"A{idx}"].border = total_border
        ws_is[f"B{idx}"].border = total_border
        ws_is[f"C{idx}"].border = total_border

# Thêm ghi chú của kế toán trưởng ở cột D
ws_is["D4"] = "Doanh thu tăng trưởng mạnh do đợt khuyến mãi hè."
ws_is["D8"] = "Tăng lương thưởng cho bộ phận Sale hoàn thành KPI."
ws_is["D10"] = "Chi phí tiếp tiếp khách đoàn khảo sát phía Nam (Chưa lấy được hóa đơn đỏ từ nhà hàng)."
ws_is["D10"].font = Font(italic=True, color="FF0000") # Bôi đỏ ghi chú nhạy cảm!

for idx in range(4, 16):
    ws_is[f"D{idx}"].border = cell_border
    if not ws_is[f"D{idx}"].value:
        ws_is[f"D{idx}"] = "-"
        ws_is[f"D{idx}"].alignment = align_center

# --- TỰ ĐỘNG CÂN CHỈNH ĐỘ RỘNG CỘT ---
for ws in [ws_assumptions, ws_is]:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or "")
            if len(val_str) > max_len:
                max_len = len(val_str)
        # Giới hạn độ rộng tối đa để giao diện cân đối
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 45)

# Lưu tệp tin
os.makedirs("D:/TaxSentry", exist_ok=True)
output_path = "D:/TaxSentry/mock_report.xlsx"
wb.save(output_path)
print(f"🎉 Đã chế tạo thành công file Excel giả lập tại: {output_path}")
