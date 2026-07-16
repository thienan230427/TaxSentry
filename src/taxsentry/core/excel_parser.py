import json
import math
import re
import unicodedata
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter

from taxsentry.config import APP_HOME

EXCEL_PATH = APP_HOME / "mock_report.xlsx"
JSON_PATH = APP_HOME / "parsed_report.json"

CANONICAL_LABEL_KEYWORDS = {
    "revenue": ["doanh thu", "doanh thu thuan", "sales", "revenue"],
    "cogs": ["gia von", "cost of goods sold", "cogs"],
    "gross_profit": ["loi nhuan gop", "gross profit"],
    "marketing_exp": ["chi phi ban hang", "marketing", "van chuyen"],
    "salary_exp": ["chi phi luong", "luong nhan vien", "salary expense", "payroll expense"],
    "hospitality_valid_exp": ["tiep khach hop le", "hospitality valid", "tiep khach co hoa don"],
    "hospitality_no_invoice_exp": ["khong co hoa don", "khong hoa don do", "hospitality no invoice", "chi phi khong hoa don"],
    "rent_exp": ["thue van phong", "rent expense", "chi phi thue"],
    "total_opex": ["tong chi phi", "tong chi phi quan ly", "tong chi phi van hanh", "opex", "total operating expenses"],
    "ebt": ["loi nhuan truoc thue", "ebt", "profit before tax"],
    "tax_expense": ["thue tndn", "chi phi thue", "tax expense", "thue phai nop"],
    "net_income": ["loi nhuan sau thue", "loi nhuan rong", "net income", "profit after tax"],
    "total_income": ["tong thu nhap", "tong tien luong", "tong quy luong", "tong quỹ lương", "gross payroll", "tong cong"],
    "personal_income_tax": ["thue tncn", "personal income tax", "pit"],
    "social_insurance": ["bhxh", "bao hiem xa hoi", "social insurance", "bhxh/bhyt/bhtn"],
    "health_insurance": ["bhyt", "bao hiem y te", "health insurance"],
    "unemployment_insurance": ["bhtn", "bao hiem that nghiep", "unemployment insurance"],
    "net_pay": ["thuc linh", "net pay"],
    "employee_count": ["so luong", "nhan vien", "employee count", "headcount"],
}

FIELD_ALIASES = {
    "revenue": "revenue",
    "cogs": "cogs",
    "gross_profit": "gross_profit",
    "marketing_exp": "marketing_exp",
    "salary_exp": "salary_exp",
    "hospitality_valid_exp": "hospitality_valid_exp",
    "hospitality_no_invoice_exp": "hospitality_no_invoice_exp",
    "rent_exp": "rent_exp",
    "total_opex": "total_opex",
    "ebt": "ebt",
    "tax_expense": "tax_expense",
    "net_income": "net_income",
}


class TaxSentryParser:
    """Bộ đọc và chuẩn hóa báo cáo Excel linh hoạt cho nhiều cấu trúc workbook."""

    def __init__(self, file_path: str):
        self.file_path = Path(file_path)
        self.wb = None
        self.wb_values = None
        self.assumptions = {}
        self.is_data = {"T4": {}, "T5": {}, "notes": {}}
        self.sheet_reports = []
        self.document_types = []
        self.canonical_metrics = {}
        self.analysis = {}
        self._analysis_done = False
        self._formula_cache = {}

    def load(self):
        """Nạp workbook ở cả chế độ raw formula và cached values."""
        if not self.file_path.exists():
            raise FileNotFoundError(f"Không tìm thấy file: {self.file_path}")

        self.wb = load_workbook(self.file_path, data_only=False)
        try:
            self.wb_values = load_workbook(self.file_path, data_only=True)
        except Exception:
            self.wb_values = None

    def parse_assumptions(self):
        self._ensure_analysis()
        return self.assumptions

    def parse_income_statement(self):
        self._ensure_analysis()
        return self.is_data

    def parse_workbook(self):
        self._ensure_analysis()
        return self.analysis

    def has_meaningful_data(self) -> bool:
        self._ensure_analysis()
        if self.canonical_metrics:
            return True
        for report in self.sheet_reports:
            if report.get("line_items") or report.get("records") or report.get("summary") or report.get("assumptions"):
                return True
        return False

    def export_json(
        self,
        output_path: str = None,
        trace_context: dict | None = None,
        artifact_store=None,
    ) -> str:
        """Xuất kết quả phân tích sạch sẽ ra định dạng JSON."""
        self._ensure_analysis()

        provenance = {
            "session_id": (trace_context or {}).get("session_id"),
            "event_id": (trace_context or {}).get("event_id"),
            "trace_id": (trace_context or {}).get("trace_id"),
            "source_file": self.file_path.name,
            "source_path": str(self.file_path),
        }
        result = {
            "metadata": {
                "project": "TaxSentry",
                "parsed_at": datetime_now_str(),
                "file_name": self.file_path.name,
                "sheet_count": len(self.wb.sheetnames) if self.wb else 0,
                "sheet_names": self.wb.sheetnames if self.wb else [],
                "document_types": self.document_types,
                "provenance": provenance,
            },
            "assumptions": self.assumptions,
            "data": {
                "income_statement": {
                    "T4_Actual": self.is_data["T4"],
                    "T5_Actual": self.is_data["T5"],
                    "notes": self.is_data["notes"],
                },
                "canonical_metrics": self.canonical_metrics,
                "workbook_overview": {
                    "document_types": self.document_types,
                    "sheet_count": len(self.wb.sheetnames) if self.wb else 0,
                    "sheet_names": self.wb.sheetnames if self.wb else [],
                },
                "sheets": self.sheet_reports,
            },
        }
        json_str = json.dumps(result, indent=4, ensure_ascii=False)
        if output_path:
            Path(output_path).write_text(json_str, encoding="utf-8")
            if artifact_store is not None:
                artifact_store.register_artifact(
                    artifact_type="parsed_json",
                    artifact_name=Path(output_path).name,
                    artifact_path=output_path,
                    session_id=provenance.get("session_id"),
                    event_id=provenance.get("event_id"),
                    trace_id=provenance.get("trace_id"),
                    source_file=self.file_path.name,
                    source_path=str(self.file_path),
                    mime_type="application/json",
                    metadata={"kind": "excel_export", "document_types": self.document_types, "provenance": provenance},
                )
        return json_str

    def log_to_database(self, trace_context: dict | None = None, job_id: str | None = None) -> bool:
        """Ghi nhận báo cáo đã phân tích vào SQLite Database."""
        import os
        from datetime import datetime

        from taxsentry.database.db_manager import TaxSentryDBManager

        self._ensure_analysis()

        db = TaxSentryDBManager()
        if not db.connect():
            return False

        metrics = self.canonical_metrics
        revenue = metrics.get("revenue", {}).get("value")
        gross_profit = metrics.get("gross_profit", {}).get("value")
        total_opex = metrics.get("total_opex", {}).get("value")
        net_income = metrics.get("net_income", {}).get("value")
        no_invoice = metrics.get("hospitality_no_invoice_exp", {}).get("value") or 0
        valid_hosp = metrics.get("hospitality_valid_exp", {}).get("value") or 0
        total_income = metrics.get("total_income", {}).get("value")
        pit = metrics.get("personal_income_tax", {}).get("value")
        social_insurance = metrics.get("social_insurance", {}).get("value")
        document_types = ", ".join(self.document_types) or "unknown"

        hospitality_limit_pct = self.assumptions.get("hospitality_limit_pct") or 0
        limit = revenue * hospitality_limit_pct if revenue and hospitality_limit_pct else 0
        total_hosp = (valid_hosp or 0) + (no_invoice or 0)

        risks = []
        if no_invoice > 0:
            risks.append("Chi phí không hóa đơn")
        if total_hosp and limit and total_hosp > limit:
            risks.append("Vượt hạn mức tiếp khách")
        if pit and pit > 0:
            risks.append("Cần rà soát Thuế TNCN")
        if social_insurance and social_insurance > 0:
            risks.append("Cần rà soát BHXH/BHYT/BHTN")

        if risks:
            tax_risk_status = "🚨 Rủi ro (" + " & ".join(sorted(set(risks))) + ")"
        else:
            tax_risk_status = f"✅ Đã phân tích ({document_types})"

        if total_opex is None and total_income is not None:
            total_opex = total_income
        if net_income is None and metrics.get("net_pay"):
            net_income = metrics["net_pay"].get("value")

        sender_email = os.getenv("ACCOUNTANT_EMAIL", "")

        success = db.log_report(
            received_at=datetime.now(),
            sender=sender_email,
            file_name=self.file_path.name,
            revenue=revenue,
            gross_profit=gross_profit,
            total_opex=total_opex,
            net_income=net_income,
            hospitality_no_invoice=no_invoice,
            tax_risk_status=tax_risk_status,
            status="Processed",
            job_id=job_id,
            session_id=(trace_context or {}).get("session_id"),
            event_id=(trace_context or {}).get("event_id"),
            trace_id=(trace_context or {}).get("trace_id"),
            source_path=str(self.file_path),
            source_file=self.file_path.name,
            trace_generated_at=(trace_context or {}).get("generated_at"),
        )
        db.close()
        return success

    def _ensure_analysis(self):
        if self._analysis_done:
            return
        if not self.wb:
            self.load()

        reports = []
        for ws in self.wb.worksheets:
            reports.append(self._analyze_sheet(ws))

        self.sheet_reports = reports
        self.document_types = list(dict.fromkeys([r["type"] for r in reports if r.get("type")]))
        self.assumptions = self._collect_assumptions(reports)
        self.canonical_metrics = self._derive_canonical_metrics(reports)
        self._populate_income_statement_fallback(reports)
        self.analysis = {
            "document_types": self.document_types,
            "sheet_reports": self.sheet_reports,
            "canonical_metrics": self.canonical_metrics,
            "assumptions": self.assumptions,
        }
        self._analysis_done = True

    def _analyze_sheet(self, ws):
        sheet_type = self._detect_sheet_type(ws)
        generic = self._extract_generic_table(ws, sheet_type)

        report = {
            "name": ws.title,
            "type": sheet_type,
            "dimensions": {"rows": ws.max_row, "cols": ws.max_column},
            "headers": generic.get("headers", []),
            "line_items": generic.get("line_items", []),
            "summary": generic.get("summary", {}),
            "notes": generic.get("notes", []),
        }

        if sheet_type == "assumptions":
            report["assumptions"] = self._extract_assumptions_from_sheet(ws, generic)
        if sheet_type == "payroll":
            report["records"] = self._extract_payroll_records(ws, generic)
            report["summary"].update(self._extract_payroll_summary(ws, generic, report.get("records", [])))
        elif sheet_type == "tax_summary":
            report["summary"].update(self._extract_tax_summary(generic))
        else:
            report["records"] = []

        if "assumptions" not in report:
            report["assumptions"] = {}

        return report

    def _detect_sheet_type(self, ws) -> str:
        title = self._normalize_text(ws.title)
        preview_texts = []
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 12), max_col=min(ws.max_column, 8), values_only=True):
            for value in row:
                text = self._normalize_text(value)
                if text:
                    preview_texts.append(text)
        corpus = " | ".join([title] + preview_texts)

        rules = [
            ("assumptions", ["assumptions", "gia dinh", "thue suat", "tax rate", "muc phat"]),
            ("income_statement", ["income_statement", "ket qua hoat dong kinh doanh", "loi nhuan", "gia von", "doanh thu"]),
            ("payroll", ["bang luong", "luong", "thuc linh", "nhan vien", "thu nhap", "thue tncn", "bhxh"]),
            ("tax_summary", ["tong hop thue", "bao hiem", "thue-bh", "bhxh", "bhyt", "bhtn"]),
            ("balance_sheet", ["can doi ke toan", "tai san", "nguon von", "balance sheet"]),
            ("cash_flow", ["luu chuyen tien te", "dong tien", "cash flow"]),
            ("ledger", ["so cai", "nhat ky", "ledger", "journal"]),
        ]

        for sheet_type, keywords in rules:
            if any(keyword in corpus for keyword in keywords):
                return sheet_type
        return "generic_table"

    def _extract_assumptions_from_sheet(self, ws, generic):
        assumptions = {}
        rows = generic.get("line_items", [])
        for item in rows:
            label_norm = self._normalize_text(item.get("label"))
            value = self._last_numeric_value(item)
            if value is None:
                continue
            if "thue suat" in label_norm or "tax rate" in label_norm:
                assumptions["tax_rate"] = value
            elif "han muc" in label_norm or "hospitality limit" in label_norm:
                assumptions["hospitality_limit_pct"] = value
            elif "phat cham nop" in label_norm or "daily penalty" in label_norm:
                assumptions["daily_penalty_pct"] = value

        if assumptions.keys() >= {"tax_rate", "hospitality_limit_pct", "daily_penalty_pct"}:
            return assumptions

        # fallback: quét trực tiếp toàn sheet cho các bảng assumptions không có header row rõ ràng
        for row_idx in range(1, ws.max_row + 1):
            text_cells = []
            numeric_values = []
            for col_idx in range(1, ws.max_column + 1):
                raw = ws.cell(row=row_idx, column=col_idx).value
                text = self._stringify(raw)
                num = self._resolved_cell_value(ws.title, row_idx, col_idx)
                if text:
                    text_cells.append(text)
                if isinstance(num, (int, float)):
                    numeric_values.append(num)

            if not text_cells or not numeric_values:
                continue

            label_norm = self._normalize_text(text_cells[0])
            value = numeric_values[-1]
            if "thue suat" in label_norm or "tax rate" in label_norm:
                assumptions["tax_rate"] = value
            elif "han muc" in label_norm or "hospitality limit" in label_norm:
                assumptions["hospitality_limit_pct"] = value
            elif "phat cham nop" in label_norm or "daily penalty" in label_norm:
                assumptions["daily_penalty_pct"] = value

        if assumptions:
            return assumptions

        # fallback cho template cũ
        try:
            assumptions["tax_rate"] = self._coerce_number(ws["B4"].value)
            assumptions["hospitality_limit_pct"] = self._coerce_number(ws["B5"].value)
            assumptions["daily_penalty_pct"] = self._coerce_number(ws["B6"].value)
        except Exception:
            pass
        return {k: v for k, v in assumptions.items() if v is not None}

    def _extract_generic_table(self, ws, sheet_type: str):
        header_row_idx, headers_map = self._detect_header_row(ws)
        line_items = []
        notes = []
        summary = {}
        blank_run = 0

        for row_idx in range((header_row_idx or 0) + 1, ws.max_row + 1):
            row_values = [ws.cell(row=row_idx, column=col_idx).value for col_idx in range(1, ws.max_column + 1)]
            if self._row_is_blank(row_values):
                blank_run += 1
                if blank_run >= 5 and line_items:
                    break
                continue
            blank_run = 0

            item = self._row_to_line_item(ws, row_idx, headers_map)
            if not item:
                row_text = " ".join([self._stringify(v) for v in row_values if self._stringify(v)])
                if row_text:
                    notes.append(row_text)
                continue

            line_items.append(item)
            if len(line_items) >= 120:
                summary["truncated_line_items"] = True
                break

        if header_row_idx:
            summary["header_row"] = header_row_idx
        if headers_map:
            summary["header_labels"] = [headers_map[k] for k in sorted(headers_map)]

        if sheet_type in {"income_statement", "balance_sheet", "cash_flow", "generic_table"}:
            summary.update(self._extract_metric_summary_from_line_items(line_items))

        return {
            "headers": [headers_map[k] for k in sorted(headers_map)] if headers_map else [],
            "line_items": line_items,
            "summary": summary,
            "notes": notes[:40],
        }

    def _extract_payroll_records(self, ws, generic):
        header_row_idx, headers_map = self._detect_header_row(ws, preferred_keywords=["ho va ten", "họ và tên", "chuc vu", "luong", "thu nhap"])
        if not header_row_idx:
            return []

        header_norm = {col: self._normalize_text(text) for col, text in headers_map.items()}
        name_col = self._find_column(header_norm, ["ho va ten", "nhan vien", "employee"])
        position_col = self._find_column(header_norm, ["chuc vu", "position"])
        if not name_col:
            return []

        records = []
        blank_run = 0
        for row_idx in range(header_row_idx + 1, ws.max_row + 1):
            row_values = [ws.cell(row=row_idx, column=c).value for c in range(1, ws.max_column + 1)]
            if self._row_is_blank(row_values):
                blank_run += 1
                if blank_run >= 4 and records:
                    break
                continue
            blank_run = 0

            label_probe = self._normalize_text(ws.cell(row=row_idx, column=name_col).value)
            if any(token in label_probe for token in ["tong cong", "tong", "total"]):
                break
            if not label_probe or label_probe in {"stt", "ho va ten", "nhan vien"}:
                continue

            metrics = {}
            for col_idx, header_text in headers_map.items():
                if col_idx in {name_col, position_col}:
                    continue
                value = self._resolved_cell_value(ws.title, row_idx, col_idx)
                if value is None:
                    continue
                if isinstance(value, (int, float)):
                    metrics[header_text] = value

            if not metrics:
                continue

            record = {
                "row": row_idx,
                "employee_name": self._stringify(ws.cell(row=row_idx, column=name_col).value),
                "position": self._stringify(ws.cell(row=row_idx, column=position_col).value) if position_col else "",
                "metrics": metrics,
            }
            records.append(record)
            if len(records) >= 150:
                break
        return records

    def _extract_payroll_summary(self, ws, generic, records):
        header_row_idx, headers_map = self._detect_header_row(ws, preferred_keywords=["ho va ten", "luong", "thu nhap"])
        if not header_row_idx:
            return {}

        name_col = self._find_column({col: self._normalize_text(text) for col, text in headers_map.items()}, ["ho va ten", "nhan vien", "employee"])
        totals = {}
        if name_col:
            for row_idx in range(header_row_idx + 1, ws.max_row + 1):
                name_text = self._normalize_text(ws.cell(row=row_idx, column=name_col).value)
                if any(token in name_text for token in ["tong cong", "tong", "total"]):
                    for col_idx, header_text in headers_map.items():
                        value = self._resolved_cell_value(ws.title, row_idx, col_idx)
                        if isinstance(value, (int, float)):
                            totals[header_text] = value
                    break

        if totals:
            return {"totals": totals}

        aggregate = defaultdict(float)
        for record in records:
            for key, value in record.get("metrics", {}).items():
                if isinstance(value, (int, float)):
                    aggregate[key] += value

        return {"totals": dict(aggregate), "record_count": len(records)} if aggregate else {}

    def _extract_tax_summary(self, generic):
        summary = {}
        for item in generic.get("line_items", []):
            label = self._normalize_text(item.get("label"))
            value = self._last_numeric_value(item)
            if value is None:
                continue
            if "bhxh" in label:
                summary.setdefault("totals", {})[item["label"]] = value
            elif "bhyt" in label:
                summary.setdefault("totals", {})[item["label"]] = value
            elif "bhtn" in label:
                summary.setdefault("totals", {})[item["label"]] = value
            elif "thue tncn" in label:
                summary.setdefault("totals", {})[item["label"]] = value
        return summary

    def _detect_header_row(self, ws, preferred_keywords=None):
        preferred_keywords = preferred_keywords or []
        best_score = -1
        best_row_idx = None
        best_headers = {}

        max_scan_row = min(ws.max_row, 20)
        for row_idx in range(1, max_scan_row + 1):
            texts = {}
            score = 0
            for col_idx in range(1, ws.max_column + 1):
                value = ws.cell(row=row_idx, column=col_idx).value
                text = self._stringify(value)
                norm = self._normalize_text(text)
                if not norm:
                    continue
                texts[col_idx] = text
                if any(ch.isalpha() for ch in norm):
                    score += 1
                if any(keyword in norm for keyword in ["thang", "month", "quy", "nam", "actual", "input", "ghi chu", "chi tieu", "ho va ten", "chuc vu", "thuc linh", "doanh thu"]):
                    score += 2
                if preferred_keywords and any(keyword in norm for keyword in [self._normalize_text(k) for k in preferred_keywords]):
                    score += 4
            if len(texts) >= 2 and score > best_score:
                best_score = score
                best_row_idx = row_idx
                best_headers = texts

        if best_score < 2:
            return None, {}
        return best_row_idx, best_headers

    def _row_to_line_item(self, ws, row_idx, headers_map):
        text_cells = []
        numeric_cells = []
        for col_idx in range(1, ws.max_column + 1):
            raw = ws.cell(row=row_idx, column=col_idx).value
            text = self._stringify(raw)
            num = self._resolved_cell_value(ws.title, row_idx, col_idx)
            if text:
                text_cells.append((col_idx, text))
            if isinstance(num, (int, float)):
                numeric_cells.append((col_idx, num))

        if not numeric_cells:
            return None

        label_col = None
        label_text = ""
        code = ""
        for col_idx, text in text_cells:
            norm = self._normalize_text(text)
            if self._looks_like_nonlabel(norm):
                continue
            label_col = col_idx
            label_text = text
            break

        if not label_text:
            return None

        if text_cells and text_cells[0][0] < (label_col or 999):
            code_candidate = self._stringify(text_cells[0][1])
            if code_candidate != label_text:
                code = code_candidate

        values = {}
        note_texts = []
        last_numeric_col = 0
        for col_idx, value in numeric_cells:
            if label_col and col_idx <= label_col:
                continue
            header = headers_map.get(col_idx) or f"Cột {get_column_letter(col_idx)}"
            values[header] = value
            last_numeric_col = max(last_numeric_col, col_idx)

        if not values:
            return None

        for col_idx, text in text_cells:
            if col_idx > last_numeric_col:
                note_texts.append(text)

        item = {
            "row": row_idx,
            "label": label_text,
            "code": code,
            "values": values,
        }
        if note_texts:
            item["note"] = " | ".join(note_texts[:3])
        return item

    def _extract_metric_summary_from_line_items(self, line_items):
        summary = {}
        for item in line_items:
            matched_key = self._match_canonical_label(item.get("label"))
            if not matched_key:
                continue
            summary[matched_key] = {
                "label": item.get("label"),
                "value": self._preferred_numeric_value(item),
            }
        return summary

    def _collect_assumptions(self, reports):
        assumptions = {}
        for report in reports:
            assumptions.update(report.get("assumptions", {}))
        return assumptions

    def _derive_canonical_metrics(self, reports):
        candidates = defaultdict(list)

        for report in reports:
            sheet_type = report.get("type")

            for item in report.get("line_items", []):
                matched_key = self._match_canonical_label(item.get("label"))
                if not matched_key:
                    continue
                value = self._preferred_numeric_value(item)
                if value is None:
                    continue
                norm = self._normalize_text(item.get("label"))
                candidates[matched_key].append({
                    "value": value,
                    "sheet": report.get("name"),
                    "sheet_type": sheet_type,
                    "label": item.get("label"),
                    "score": self._candidate_score(sheet_type, item.get("label"))
                    + (2 if matched_key == "revenue" and "doanh thu thuan" in norm else 0),
                    "periods": list(item.get("values", {}).keys()),
                })

            totals = report.get("summary", {}).get("totals", {})
            for label, value in totals.items():
                matched_key = self._match_canonical_label(label)
                if matched_key and isinstance(value, (int, float)):
                    candidates[matched_key].append({
                        "value": value,
                        "sheet": report.get("name"),
                        "sheet_type": sheet_type,
                        "label": label,
                        "score": self._candidate_score(sheet_type, label) + 1,
                        "periods": [label],
                    })

        canonical = {}
        for key, bucket in candidates.items():
            best = max(bucket, key=lambda x: x["score"])
            canonical[key] = {
                "value": best["value"],
                "source_sheet": best["sheet"],
                "source_type": best["sheet_type"],
                "source_label": best["label"],
                "periods": best.get("periods", []),
            }
        return canonical

    def _populate_income_statement_fallback(self, reports):
        self.is_data = {"T4": {}, "T5": {}, "notes": {}}

        best_report = None
        for report in reports:
            if report.get("type") == "income_statement" and report.get("line_items"):
                best_report = report
                break
        if not best_report:
            for report in reports:
                if report.get("line_items") and any(self._match_canonical_label(item.get("label")) for item in report.get("line_items", [])):
                    best_report = report
                    break
        if not best_report:
            return

        for item in best_report.get("line_items", []):
            canonical = self._match_canonical_label(item.get("label"))
            if canonical not in FIELD_ALIASES:
                continue
            values_by_period = [
                (self._normalize_text(period), value)
                for period, value in item.get("values", {}).items()
                if isinstance(value, (int, float)) and not self._is_metadata_header(period)
            ]
            if any(period in {"ky nay", "ky truoc", "current", "previous"} for period, _ in values_by_period):
                continue
            values = [value for _, value in values_by_period]
            alias = FIELD_ALIASES[canonical]
            if values:
                self.is_data["T5"][alias] = values[-1]
                if len(values) >= 2:
                    self.is_data["T4"][alias] = values[0]
            if item.get("note"):
                self.is_data["notes"][alias] = item.get("note")

    def _match_canonical_label(self, label):
        norm = self._normalize_text(label)
        if not norm:
            return None
        if "doanh thu hoat dong tai chinh" in norm or "financial income" in norm:
            return None
        for canonical, keywords in CANONICAL_LABEL_KEYWORDS.items():
            if any(keyword in norm for keyword in keywords):
                return canonical
        return None

    def _candidate_score(self, sheet_type: str, label: str) -> int:
        score = 1
        if sheet_type == "income_statement":
            score += 5
        elif sheet_type == "payroll":
            score += 4
        elif sheet_type == "tax_summary":
            score += 3
        elif sheet_type in {"balance_sheet", "cash_flow"}:
            score += 2
        norm = self._normalize_text(label)
        if any(token in norm for token in ["tong", "total", "thuc te", "actual", "thang", "quy"]):
            score += 1
        return score

    def _last_numeric_value(self, item):
        values = [v for v in item.get("values", {}).values() if isinstance(v, (int, float))]
        return values[-1] if values else None

    def _preferred_numeric_value(self, item):
        values = item.get("values", {})
        for period, value in values.items():
            if self._normalize_text(period) in {"ky nay", "current", "this period"} and isinstance(value, (int, float)):
                return value
        candidates = [
            value
            for period, value in values.items()
            if isinstance(value, (int, float)) and not self._is_metadata_header(period)
        ]
        return candidates[-1] if candidates else None

    def _is_metadata_header(self, header):
        return self._normalize_text(header) in {"ma so", "code", "stt", "thuyet minh", "note", "notes"}

    def _resolved_cell_value(self, sheet_name: str, row_idx: int, col_idx: int, visited=None):
        cell_ref = f"{get_column_letter(col_idx)}{row_idx}"
        cache_key = (sheet_name, cell_ref)
        if cache_key in self._formula_cache:
            return self._formula_cache[cache_key]

        visited = visited or set()
        if cache_key in visited:
            return 0.0
        visited.add(cache_key)

        if self.wb_values and sheet_name in self.wb_values.sheetnames:
            cached = self.wb_values[sheet_name][cell_ref].value
            cached_num = self._coerce_number(cached)
            if cached_num is not None:
                self._formula_cache[cache_key] = cached_num
                return cached_num

        value = self.wb[sheet_name][cell_ref].value
        numeric = self._coerce_number(value)
        if numeric is not None:
            self._formula_cache[cache_key] = numeric
            return numeric

        if isinstance(value, str) and value.startswith("="):
            numeric = self._evaluate_formula(sheet_name, value, visited=visited)
            self._formula_cache[cache_key] = numeric
            return numeric

        return None

    def _evaluate_formula(self, current_sheet: str, formula: str, visited=None) -> float:
        visited = visited or set()
        expr = formula.lstrip("=").replace("$", "")
        expr = self._replace_aggregate_functions(expr, current_sheet, visited)
        expr = self._replace_cross_sheet_refs(expr, current_sheet, visited)
        expr = self._replace_local_refs(expr, current_sheet, visited)
        expr = expr.replace("^", "**")
        if not re.fullmatch(r"[0-9eE\.\+\-\*\/\(\)\s\*]+", expr):
            return 0.0
        try:
            result = eval(expr, {"__builtins__": {}}, {})
            return float(result)
        except Exception:
            return 0.0

    def _replace_aggregate_functions(self, expr: str, current_sheet: str, visited) -> str:
        pattern = re.compile(r"(SUM|AVERAGE|MIN|MAX)\(([^()]+)\)", re.IGNORECASE)
        while True:
            match = pattern.search(expr)
            if not match:
                break
            func = match.group(1).upper()
            arg_text = match.group(2)
            values = []
            for token in [part.strip() for part in arg_text.split(",") if part.strip()]:
                if ":" in token:
                    if "!" in token:
                        sheet_part, range_part = token.split("!", 1)
                        sheet_name = self._strip_sheet_quotes(sheet_part)
                    else:
                        sheet_name = current_sheet
                        range_part = token
                    start_ref, end_ref = range_part.split(":", 1)
                    values.extend(self._range_values(sheet_name, start_ref, end_ref, visited))
                else:
                    values.append(self._token_to_number(token, current_sheet, visited))
            values = [v for v in values if isinstance(v, (int, float))]
            if not values:
                replacement = "0"
            elif func == "SUM":
                replacement = str(sum(values))
            elif func == "AVERAGE":
                replacement = str(sum(values) / len(values))
            elif func == "MIN":
                replacement = str(min(values))
            else:
                replacement = str(max(values))
            expr = expr[:match.start()] + replacement + expr[match.end():]
        return expr

    def _replace_cross_sheet_refs(self, expr: str, current_sheet: str, visited) -> str:
        pattern = re.compile(r"((?:'[^']+'|[A-Za-z0-9_\- ]+))!([A-Z]{1,3}\d+)", re.IGNORECASE)
        def repl(match):
            sheet_name = self._strip_sheet_quotes(match.group(1))
            cell_ref = match.group(2)
            col_letters = re.sub(r"\d+", "", cell_ref)
            row_number = int(re.sub(r"[A-Z]+", "", cell_ref, flags=re.IGNORECASE))
            value = self._resolved_cell_value(sheet_name, row_number, column_index_from_string(col_letters), visited=visited.copy())
            return str(value or 0)
        return pattern.sub(repl, expr)

    def _replace_local_refs(self, expr: str, current_sheet: str, visited) -> str:
        pattern = re.compile(r"(?<![A-Z0-9_])([A-Z]{1,3}\d+)(?![A-Z0-9_])", re.IGNORECASE)
        def repl(match):
            cell_ref = match.group(1)
            col_letters = re.sub(r"\d+", "", cell_ref)
            row_number = int(re.sub(r"[A-Z]+", "", cell_ref, flags=re.IGNORECASE))
            value = self._resolved_cell_value(current_sheet, row_number, column_index_from_string(col_letters), visited=visited.copy())
            return str(value or 0)
        return pattern.sub(repl, expr)

    def _range_values(self, sheet_name: str, start_ref: str, end_ref: str, visited):
        start_col = column_index_from_string(re.sub(r"\d+", "", start_ref, flags=re.IGNORECASE))
        start_row = int(re.sub(r"[A-Z]+", "", start_ref, flags=re.IGNORECASE))
        end_col = column_index_from_string(re.sub(r"\d+", "", end_ref, flags=re.IGNORECASE))
        end_row = int(re.sub(r"[A-Z]+", "", end_ref, flags=re.IGNORECASE))
        values = []
        for row_idx in range(min(start_row, end_row), max(start_row, end_row) + 1):
            for col_idx in range(min(start_col, end_col), max(start_col, end_col) + 1):
                value = self._resolved_cell_value(sheet_name, row_idx, col_idx, visited=visited.copy())
                if isinstance(value, (int, float)):
                    values.append(value)
        return values

    def _token_to_number(self, token: str, current_sheet: str, visited):
        token = token.strip()
        direct = self._coerce_number(token)
        if direct is not None:
            return direct
        if "!" in token:
            sheet_part, cell_ref = token.split("!", 1)
            sheet_name = self._strip_sheet_quotes(sheet_part)
        else:
            sheet_name = current_sheet
            cell_ref = token
        cell_ref = cell_ref.replace("$", "")
        if not re.fullmatch(r"[A-Z]{1,3}\d+", cell_ref, re.IGNORECASE):
            return 0.0
        col_letters = re.sub(r"\d+", "", cell_ref, flags=re.IGNORECASE)
        row_number = int(re.sub(r"[A-Z]+", "", cell_ref, flags=re.IGNORECASE))
        return self._resolved_cell_value(sheet_name, row_number, column_index_from_string(col_letters), visited=visited.copy()) or 0.0

    def _find_column(self, header_norm_map, keywords):
        normalized_keywords = [self._normalize_text(k) for k in keywords]
        for col_idx, header_norm in header_norm_map.items():
            if any(keyword in header_norm for keyword in normalized_keywords):
                return col_idx
        return None

    def _looks_like_nonlabel(self, norm_text: str) -> bool:
        if not norm_text:
            return True
        if norm_text in {"stt", "no", "ma", "code", "id"}:
            return True
        if re.fullmatch(r"[0-9\-\./]+", norm_text):
            return True
        return False

    def _row_is_blank(self, row_values):
        return not any(self._stringify(v) or self._coerce_number(v) is not None for v in row_values)

    def _coerce_number(self, value: Any):
        if value is None or isinstance(value, bool):
            return None
        if isinstance(value, (int, float)):
            if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
                return None
            return float(value)
        if isinstance(value, str):
            text = value.strip()
            if not text:
                return None
            text = text.replace("₫", "").replace("VND", "").replace("vnd", "")
            text = text.replace(",", "")
            if text.startswith("(") and text.endswith(")"):
                text = "-" + text[1:-1]
            if text.endswith("%"):
                try:
                    return float(text[:-1]) / 100.0
                except Exception:
                    return None
            if re.fullmatch(r"-?\d+(\.\d+)?", text):
                try:
                    return float(text)
                except Exception:
                    return None
        return None

    def _normalize_text(self, value: Any) -> str:
        text = self._stringify(value).lower()
        if not text:
            return ""
        text = unicodedata.normalize("NFKD", text)
        text = "".join(ch for ch in text if not unicodedata.combining(ch))
        text = text.replace("đ", "d")
        text = re.sub(r"\s+", " ", text)
        return text.strip()

    def _stringify(self, value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, str):
            return value.strip()
        return str(value).strip()

    def _strip_sheet_quotes(self, sheet_name: str) -> str:
        sheet_name = sheet_name.strip()
        if sheet_name.startswith("'") and sheet_name.endswith("'"):
            return sheet_name[1:-1]
        return sheet_name


def datetime_now_str():
    from datetime import datetime
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def main():
    print("--- CHẠY THỬ NGHIỆM TAXSENTRY PARSER ---")
    input_file = str(EXCEL_PATH)
    output_json = str(JSON_PATH)

    parser = TaxSentryParser(input_file)
    try:
        parser.load()
        parser.parse_workbook()

        json_output = parser.export_json(output_json)
        print(f"✅ Đã trích xuất và chuẩn hóa dữ liệu thành công ra: {output_json}")
        print(f"📚 Loại tài liệu phát hiện: {', '.join(parser.document_types) or 'unknown'}")
        print(f"📊 Metrics nhận diện: {', '.join(parser.canonical_metrics.keys()) or 'không có'}")

        print("Đang tự động đồng bộ kết quả phân tích vào Database...")
        if parser.log_to_database():
            print("✅ Đã ghi nhận và lưu trữ báo cáo thành công vào Database 'tax_sentry'!")
        else:
            print("❌ Lưu trữ Database thất bại.")
        print()
        data = json.loads(json_output)
        print(json.dumps({
            "document_types": data["metadata"]["document_types"],
            "canonical_metrics": data["data"]["canonical_metrics"],
            "income_statement_summary_T5": data["data"]["income_statement"]["T5_Actual"],
        }, indent=4, ensure_ascii=False))
    except Exception as e:
        print(f"❌ Có lỗi trong quá trình đọc file: {e}")


if __name__ == "__main__":
    main()
