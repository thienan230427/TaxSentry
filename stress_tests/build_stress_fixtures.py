from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill

base = Path(r'D:\TaxSentry\stress_tests')
base.mkdir(parents=True, exist_ok=True)

BLUE = Font(color='0000FF')
BOLD = Font(bold=True)
WHITE_BOLD = Font(color='FFFFFF', bold=True)
HEADER_FILL = PatternFill('solid', fgColor='1F4E79')
SUB_FILL = PatternFill('solid', fgColor='D9EAF7')

# 1) complex financial pack
wb = Workbook()
ws = wb.active
ws.title = 'Assumptions'
ws['A1'] = 'BẢNG GIẢ ĐỊNH THUẾ & RỦI RO'
ws['A1'].font = WHITE_BOLD
ws['A1'].fill = HEADER_FILL
ws.merge_cells('A1:C1')
for c, h in enumerate(['Chỉ số', 'Giá trị', 'Nguồn'], 1):
    cell = ws.cell(row=3, column=c, value=h)
    cell.font = BOLD
    cell.fill = SUB_FILL
rows = [
    ('Thuế suất Thuế TNDN', 0.20, 'Luật thuế TNDN hiện hành'),
    ('Hạn mức chi phí Tiếp khách hợp lệ (% Doanh thu)', 0.15, 'Quy chế nội bộ + luật thuế'),
    ('Mức phạt chậm nộp thuế mỗi ngày', 0.0003, 'Luật quản lý thuế'),
]
for r, row in enumerate(rows, 4):
    for c, v in enumerate(row, 1):
        cell = ws.cell(row=r, column=c, value=v)
        if c == 2:
            cell.font = BLUE
            cell.comment = Comment('Source: fixture stress test', 'Grace')

is_ws = wb.create_sheet('Income_Statement')
is_ws['A1'] = 'BÁO CÁO KẾT QUẢ HOẠT ĐỘNG KINH DOANH — NEOFIN PREMIUM'
is_ws['A1'].font = WHITE_BOLD
is_ws['A1'].fill = HEADER_FILL
is_ws.merge_cells('A1:D1')
headers = ['Chỉ tiêu kinh doanh (VND)', 'Quý I/2026 (Thực tế)', 'Quý II/2026 (Nhập liệu)', 'Ghi chú']
for c, h in enumerate(headers, 1):
    cell = is_ws.cell(row=4, column=c, value=h)
    cell.font = BOLD
    cell.fill = SUB_FILL
items = [
    ('Doanh thu bán hàng và cung cấp dịch vụ', 1200000000, 1580000000, 'Mở rộng kênh B2B'),
    ('Giá vốn hàng bán', 420000000, 540000000, 'Chi phí cloud + vendor'),
    ('Lợi nhuận gộp', '=B5-B6', '=C5-C6', '-'),
    ('Chi phí bán hàng (Marketing, vận chuyển)', 180000000, 260000000, 'Chiến dịch AppSumo thử nghiệm'),
    ('Chi phí lương nhân viên', 260000000, 330000000, 'Tuyển thêm đội growth & AI'),
    ('Chi phí tiếp khách (Hợp lệ có hóa đơn)', 45000000, 58000000, 'Tiếp đối tác'),
    ('Chi phí tiếp khách (KHÔNG có hóa đơn đỏ)', 12000000, 68000000, 'Dinner founder circle, thiếu chứng từ'),
    ('Chi phí thuê văn phòng và dịch vụ khác', 90000000, 110000000, 'Office + SaaS'),
    ('Tổng chi phí quản lý & vận hành', '=SUM(B8:B12)', '=SUM(C8:C12)', '-'),
    ('Lợi nhuận kế toán trước thuế', '=B7-B13', '=C7-C13', '-'),
    ('Thuế TNDN phải nộp', '=B14*Assumptions!$B$4', '=C14*Assumptions!$B$4', '-'),
    ('Lợi nhuận ròng sau thuế', '=B14-B15', '=C14-C15', '-'),
]
for r, row in enumerate(items, 5):
    for c, v in enumerate(row, 1):
        cell = is_ws.cell(row=r, column=c, value=v)
        if c in (2, 3) and not str(v).startswith('='):
            cell.font = BLUE
        if c in (2, 3):
            cell.number_format = '#,##0'

bs = wb.create_sheet('Balance Sheet 2026')
bs['A1'] = 'BẢNG CÂN ĐỐI KẾ TOÁN TÓM TẮT'
bs['A1'].font = WHITE_BOLD
bs['A1'].fill = HEADER_FILL
bs.merge_cells('A1:E1')
for c, h in enumerate(['Mã', 'Chỉ tiêu', '31/03/2026', '30/06/2026', 'Ghi chú'], 1):
    cell = bs.cell(row=3, column=c, value=h)
    cell.font = BOLD
    cell.fill = SUB_FILL
bs_rows = [
    ('100', 'Tiền và tương đương tiền', 260000000, 310000000, ''),
    ('110', 'Các khoản phải thu khách hàng', 145000000, 182000000, ''),
    ('120', 'Hàng tồn kho', 87000000, 93000000, ''),
    ('300', 'Phải trả người bán', 76000000, 81000000, ''),
    ('400', 'Vốn chủ sở hữu', 416000000, 504000000, ''),
]
for r, row in enumerate(bs_rows, 4):
    for c, v in enumerate(row, 1):
        bs.cell(row=r, column=c, value=v)

cf = wb.create_sheet('CashFlow Mix')
cf['A1'] = 'BÁO CÁO LƯU CHUYỂN TIỀN TỆ — GIÁN TIẾP'
cf['A1'].font = WHITE_BOLD
cf['A1'].fill = HEADER_FILL
cf.merge_cells('A1:C1')
for c, h in enumerate(['Chỉ tiêu', 'Q1', 'Q2'], 1):
    cell = cf.cell(row=3, column=c, value=h)
    cell.font = BOLD
    cell.fill = SUB_FILL
cf['A4'] = 'Lưu chuyển tiền từ HĐKD'; cf['B4'] = 180000000; cf['C4'] = 240000000
cf['A5'] = 'Lưu chuyển tiền từ HĐĐT'; cf['B5'] = -65000000; cf['C5'] = -82000000
cf['A6'] = 'Lưu chuyển tiền từ HĐTC'; cf['B6'] = 30000000; cf['C6'] = 40000000
wb.save(base / 'stress_financial_multisheet.xlsx')

# 2) payroll complex
wb = Workbook()
ws = wb.active
ws.title = 'Bảng lương T07-2026'
ws.merge_cells('A1:N1')
ws['A1'] = 'BẢNG LƯƠNG THÁNG 07/2026 — FINAI HOLDINGS'
ws['A1'].font = Font(bold=True, size=16)
ws['A2'] = 'Báo cáo nội bộ | dữ liệu gồm lương, thưởng, bảo hiểm, thuế TNCN'
headers = ['STT', 'Họ và tên', 'Chức vụ', 'Lương cơ bản', 'Allowance', 'Bonus KPI', 'Commission', 'OT', 'Tổng thu nhập', 'BHXH/BHYT/BHTN', 'Thuế TNCN', 'Khấu trừ khác', 'Thực lĩnh', 'Ghi chú']
for c, h in enumerate(headers, 1):
    cell = ws.cell(row=5, column=c, value=h)
    cell.font = BOLD
    cell.fill = SUB_FILL
people = [
    ('Nguyễn Hữu Phúc', 'CEO', 90000000, 15000000, 22000000, 5000000, 3000000, 0, 'Founder'),
    ('Trần Gia Hân', 'CFO', 62000000, 10000000, 18000000, 2000000, 1500000, 0, 'Finance'),
    ('Lê Minh Tâm', 'Engineering Manager', 48000000, 8000000, 14000000, 1000000, 2000000, 500000, 'Platform'),
    ('Phạm Mỹ Linh', 'HRBP', 25000000, 3500000, 5000000, 0, 1000000, 500000, 'People Ops'),
    ('Đỗ Tuấn Kiệt', 'Accountant', 22000000, 2500000, 4000000, 0, 500000, 250000, 'Tax filing'),
]
for idx, row in enumerate(people, 6):
    name, pos, base_pay, allowance, bonus, commission, ot, deduction, note = row
    gross_formula = f'=D{idx}+E{idx}+F{idx}+G{idx}+H{idx}'
    ins_formula = f'=MIN(D{idx},36800000)*10.5%'
    tax_formula = f'=MAX(I{idx}-J{idx}-11000000,0)*10%'
    net_formula = f'=I{idx}-J{idx}-K{idx}-L{idx}'
    vals = [idx - 5, name, pos, base_pay, allowance, bonus, commission, ot, gross_formula, ins_formula, tax_formula, deduction, net_formula, note]
    for c, v in enumerate(vals, 1):
        cell = ws.cell(row=idx, column=c, value=v)
        if c in [4, 5, 6, 7, 8, 12] and not str(v).startswith('='):
            cell.font = BLUE
        if 4 <= c <= 13:
            cell.number_format = '#,##0'
tr = 11
ws.merge_cells(f'A{tr}:C{tr}')
ws[f'A{tr}'] = 'TỔNG CỘNG'
for c in range(4, 14):
    col = chr(64 + c)
    ws.cell(row=tr, column=c, value=f'=SUM({col}6:{col}10)')

summary = wb.create_sheet('Tổng hợp thuế-BH')
summary['A1'] = 'TỔNG HỢP THUẾ TNCN & BẢO HIỂM'
summary['A1'].font = Font(bold=True, size=14)
for c, h in enumerate(['STT', 'Chỉ tiêu', 'Số lượng', 'Tổng tiền', 'BHXH', 'BHYT', 'BHTN'], 1):
    cell = summary.cell(row=3, column=c, value=h)
    cell.font = BOLD
    cell.fill = SUB_FILL
summary_rows = [
    (1, 'Tổng quỹ lương tháng 07/2026', 5, "=SUM('Bảng lương T07-2026'!I6:I10)", "=SUM('Bảng lương T07-2026'!J6:J10)*0.7619", "=SUM('Bảng lương T07-2026'!J6:J10)*0.1429", "=SUM('Bảng lương T07-2026'!J6:J10)*0.0952"),
    (2, 'Thuế TNCN tạm khấu trừ', 5, "=SUM('Bảng lương T07-2026'!K6:K10)", 0, 0, 0),
]
for r, row in enumerate(summary_rows, 4):
    for c, v in enumerate(row, 1):
        summary.cell(row=r, column=c, value=v)

info = wb.create_sheet('Thông tin nhân sự mở rộng')
info['A1'] = 'Danh mục nhân sự & trạng thái HĐLĐ'
for c, h in enumerate(['Nhân viên', 'Loại hợp đồng', 'MST cá nhân', 'Ghi chú'], 1):
    cell = info.cell(row=3, column=c, value=h)
    cell.font = BOLD
    cell.fill = SUB_FILL
extra = [
    ('Nguyễn Hữu Phúc', 'Không thời hạn', '0123456789', 'Có ESOP'),
    ('Trần Gia Hân', 'Không thời hạn', '1234567890', 'Có thưởng hiệu suất'),
    ('Lê Minh Tâm', 'Không thời hạn', '2345678901', 'Quản lý trung tâm kỹ thuật'),
]
for r, row in enumerate(extra, 4):
    for c, v in enumerate(row, 1):
        info.cell(row=r, column=c, value=v)
wb.save(base / 'stress_payroll_formula_heavy.xlsx')

# 3) raw mixed single-sheet
wb = Workbook()
ws = wb.active
ws.title = 'DataRoom'
ws['A1'] = 'BỘ SỐ LIỆU TÀI CHÍNH THÔ — SHEET TỔNG HỢP'
for c, h in enumerate(['Mã mục', 'Diễn giải', 'Kỳ trước', 'Kỳ này', 'Memo'], 1):
    cell = ws.cell(row=3, column=c, value=h)
    cell.font = BOLD
    cell.fill = SUB_FILL
raw_rows = [
    ('400', 'Doanh thu thuần', 980000000, 1260000000, 'Bán gói enterprise'),
    ('410', 'Giá vốn hàng bán', 360000000, 430000000, 'Tăng do infra'),
    ('420', 'Lợi nhuận trước thuế', 190000000, 265000000, 'Chưa loại trừ một số chi phí'),
    ('421', 'Thuế TNDN phải nộp', 38000000, 53000000, 'Tạm tính 20%'),
    ('422', 'Lợi nhuận sau thuế', 152000000, 212000000, 'Theo kế toán nội bộ'),
    ('500', 'Tổng chi phí vận hành / OPEX', 240000000, 295000000, 'Bao gồm lương và S&M'),
]
for r, row in enumerate(raw_rows, 4):
    for c, v in enumerate(row, 1):
        ws.cell(row=r, column=c, value=v)
wb.save(base / 'stress_raw_single_sheet.xlsx')

# 4) hybrid weird headers
wb = Workbook()
ws = wb.active
ws.title = 'Ops-Fin Summary'
ws['B2'] = 'OPS + FIN SUMMARY'
for col, val in zip(['B4', 'C4', 'D4', 'E4'], ['Metric Name', 'Actual May', 'Forecast Jun', 'Commentary']):
    ws[col] = val
    ws[col].font = BOLD
    ws[col].fill = SUB_FILL
hybrid = [
    ('Revenue', 2100000000, 2380000000, 'Strong pipeline close'),
    ('Gross Profit', 1220000000, 1410000000, 'Cloud margin stable'),
    ('Salary Expense', 390000000, 415000000, 'New hires'),
    ('Marketing Expense', 210000000, 275000000, 'Brand launch'),
    ('Tax Expense', 115000000, 138000000, 'Need review'),
    ('Net Income', 505000000, 582000000, 'Healthy run rate'),
]
for r, row in enumerate(hybrid, 5):
    for idx, v in enumerate(row, 2):
        ws.cell(row=r, column=idx, value=v)
ws2 = wb.create_sheet('Thuế & pháp lý')
ws2['A1'] = 'Thuế & pháp lý'
for col, val in zip(['A3', 'B3', 'C3'], ['Mục', 'Giá trị', 'Ghi chú']):
    ws2[col] = val
    ws2[col].font = BOLD
    ws2[col].fill = SUB_FILL
ws2.append(['Thuế suất Thuế TNDN', 0.2, 'Chuẩn'])
ws2.append(['Mức phạt chậm nộp thuế mỗi ngày', 0.0003, 'Chuẩn'])
wb.save(base / 'stress_hybrid_bilingual.xlsx')

print(base)
for p in sorted(base.glob('*.xlsx')):
    print(p.name)
