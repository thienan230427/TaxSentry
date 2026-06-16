import json
import re
from pathlib import Path
from openpyxl import load_workbook


class TaxSentryParser:
    """Bộ đọc và chuẩn hóa báo cáo tài chính Excel (TaxSentry Parser).

    Đọc tệp Excel và tự động tính toán/giải các công thức Excel cơ bản (SUM, -, *)
    để trích xuất dữ liệu sạch dạng JSON mà không cần LibreOffice.
    """

    def __init__(self, file_path: str):
        self.file_path = Path(file_path)
        self.wb = None
        self.assumptions = {}
        self.is_data = {"T4": {}, "T5": {}, "notes": {}}

    def load(self):
        """Nạp tệp Excel."""
        if not self.file_path.exists():
            raise FileNotFoundError(f"Không tìm thấy file: {self.file_path}")
        self.wb = load_workbook(self.file_path, data_only=False)

    def parse_assumptions(self):
        """Đọc và phân tích các chỉ số giả định thuế từ tab Assumptions."""
        ws = self.wb["Assumptions"]
        # Đọc dữ liệu từ hàng 4 đến 6
        self.assumptions = {
            "tax_rate": ws["B4"].value,                      # Thuế suất TNDN (20%)
            "hospitality_limit_pct": ws["B5"].value,         # Hạn mức tiếp khách (15%)
            "daily_penalty_pct": ws["B6"].value,             # Phạt chậm nộp (0.03%)
        }

    def _resolve_formula(self, formula: str, col_letter: str, current_row: int) -> float:
        """Giải công thức toán học cơ bản một cách thông minh."""
        # Bỏ dấu '=' ở đầu
        expr = formula.lstrip("=")

        # 1. Xử lý SUM: ví dụ SUM(C7:C11)
        sum_match = re.search(r"SUM\(([A-Z]+)(\d+):([A-Z]+)(\d+)\)", expr, re.IGNORECASE)
        if sum_match:
            start_col, start_row, end_col, end_row = sum_match.groups()
            start_row, end_row = int(start_row), int(end_row)
            total_sum = 0.0
            for r in range(start_row, end_row + 1):
                cell_ref = f"{col_letter}{r}"
                total_sum += float(self._get_cell_value(col_letter, r))
            return total_sum

        # 2. Xử lý phép nhân liên kết tab Assumptions: ví dụ C13*Assumptions!$B$4
        if "Assumptions!" in expr:
            # Lấy tên ô ở Income_Statement
            local_ref = expr.split("*")[0].strip().replace("$", "")
            local_row = int(re.sub(r"[A-Z]+", "", local_ref))
            local_val = float(self._get_cell_value(col_letter, local_row))
            # Nhân với thuế suất
            return local_val * self.assumptions["tax_rate"]

        # 3. Xử lý phép trừ đơn giản: ví dụ C4-C5 hoặc C6-C12
        minus_match = re.match(r"([A-Z]+)(\d+)-([A-Z]+)(\d+)", expr, re.IGNORECASE)
        if minus_match:
            col1, r1, col2, r2 = minus_match.groups()
            val1 = float(self._get_cell_value(col1, int(r1)))
            val2 = float(self._get_cell_value(col2, int(r2)))
            return val1 - val2

        return 0.0

    def _get_cell_value(self, col_letter: str, row_idx: int) -> float:
        """Lấy giá trị thực tế của một ô (nếu là công thức thì tự giải)."""
        ws = self.wb["Income_Statement"]
        cell_ref = f"{col_letter}{row_idx}"
        val = ws[cell_ref].value

        if isinstance(val, str) and val.startswith("="):
            # Nếu là công thức, tiến hành giải công thức và lưu trữ đệm (caching)
            resolved_val = self._resolve_formula(val, col_letter, row_idx)
            ws[cell_ref].value = resolved_val  # Lưu đệm để không cần tính lại
            return resolved_val

        return float(val or 0.0)

    def parse_income_statement(self):
        """Phân tích báo cáo hoạt động kinh doanh (tab Income_Statement)."""
        ws = self.wb["Income_Statement"]

        # Các dòng chỉ tiêu tương ứng với index hàng (1-indexed trong Excel)
        rows_mapping = {
            4: "revenue",
            5: "cogs",
            6: "gross_profit",
            7: "marketing_exp",
            8: "salary_exp",
            9: "hospitality_valid_exp",
            10: "hospitality_no_invoice_exp",
            11: "rent_exp",
            12: "total_opex",
            13: "ebt",                  # Lợi nhuận trước thuế
            14: "tax_expense",          # Chi phí thuế TNDN
            15: "net_income",           # Lợi nhuận ròng sau thuế
        }

        # Đọc dữ liệu T4 (Cột B) và T5 (Cột C)
        for r_idx, name in rows_mapping.items():
            self.is_data["T4"][name] = self._get_cell_value("B", r_idx)
            self.is_data["T5"][name] = self._get_cell_value("C", r_idx)
            # Lấy ghi chú giải trình của kế toán trưởng (Cột D)
            note_val = ws[f"D{r_idx}"].value
            self.is_data["notes"][name] = note_val if note_val != "-" else None

    def export_json(self, output_path: str = None) -> str:
        """Xuất kết quả phân tích sạch sẽ ra định dạng JSON."""
        result = {
            "metadata": {
                "project": "TaxSentry",
                "parsed_at": datetime_now_str(),
                "file_name": self.file_path.name,
            },
            "assumptions": self.assumptions,
            "data": {
                "income_statement": {
                    "T4_Actual": self.is_data["T4"],
                    "T5_Actual": self.is_data["T5"],
                    "notes": self.is_data["notes"],
                }
            },
        }
        json_str = json.dumps(result, indent=4, ensure_ascii=False)
        if output_path:
            Path(output_path).write_text(json_str, encoding="utf-8")
        return json_str

    def log_to_database(self) -> bool:
        """Tự động ghi nhận báo cáo đã phân tích vào MySQL Database."""
        from db_manager import TaxSentryDBManager
        from datetime import datetime
        
        db = TaxSentryDBManager()
        if not db.connect():
            return False
            
        is_t5 = self.is_data["T5"]
        revenue = is_t5["revenue"]
        gross_profit = is_t5["gross_profit"]
        total_opex = is_t5["total_opex"]
        net_income = is_t5["net_income"]
        no_invoice = is_t5["hospitality_no_invoice_exp"]
        valid_hosp = is_t5["hospitality_valid_exp"]
        
        # Đánh giá trạng thái rủi ro thuế
        limit = revenue * self.assumptions["hospitality_limit_pct"]
        total_hosp = valid_hosp + no_invoice
        
        risks = []
        if no_invoice > 0:
            risks.append("Chi phí không hóa đơn")
        if total_hosp > limit:
            risks.append("Vượt hạn mức")
            
        if risks:
            tax_risk_status = "🚨 Rủi ro (" + " & ".join(risks) + ")"
        else:
            tax_risk_status = "✅ An toàn"
            
        success = db.log_report(
            received_at=datetime.now(), # Lấy thời điểm phân tích thực tế
            sender="thienan12342007@gmail.com", # Email Kế toán trưởng
            file_name=self.file_path.name,
            revenue=revenue,
            gross_profit=gross_profit,
            total_opex=total_opex,
            net_income=net_income,
            hospitality_no_invoice=no_invoice,
            tax_risk_status=tax_risk_status,
            status="Processed"
        )
        db.close()
        return success


def datetime_now_str():
    from datetime import datetime
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def main():
    print("--- CHẠY THỬ NGHIỆM TAXSENTRY PARSER ---")
    input_file = "D:/TaxSentry/mock_report.xlsx"
    output_json = "D:/TaxSentry/parsed_report.json"

    parser = TaxSentryParser(input_file)
    try:
        parser.load()
        parser.parse_assumptions()
        parser.parse_income_statement()
        
        # Xuất JSON
        json_output = parser.export_json(output_json)
        print(f"✅ Đã trích xuất và chuẩn hóa dữ liệu thành công ra: {output_json}")
        
        # Tự động ghi vào MySQL Database
        print("Đang tự động đồng bộ kết quả phân tích vào MySQL Database...")
        if parser.log_to_database():
            print("✅ Đã ghi nhận và lưu trữ báo cáo thành công vào Database 'tax_sentry'!")
        else:
            print("❌ Lưu trữ Database thất bại.")
        print()
        print("Dữ liệu JSON trích xuất (Trích đoạn):")
        
        # In một phần dữ liệu ra console để Sếp xem
        data = json.loads(json_output)
        print(json.dumps({
            "assumptions": data["assumptions"],
            "income_statement_summary_T5": {
                "Revenue": data["data"]["income_statement"]["T5_Actual"]["revenue"],
                "Gross Profit": data["data"]["income_statement"]["T5_Actual"]["gross_profit"],
                "Total OPEX": data["data"]["income_statement"]["T5_Actual"]["total_opex"],
                "EBT (Before Tax)": data["data"]["income_statement"]["T5_Actual"]["ebt"],
                "Tax Expense": data["data"]["income_statement"]["T5_Actual"]["tax_expense"],
                "Net Income": data["data"]["income_statement"]["T5_Actual"]["net_income"],
                "Note on Hospitality": data["data"]["income_statement"]["notes"]["hospitality_no_invoice_exp"]
            }
        }, indent=4, ensure_ascii=False))

    except Exception as e:
        print(f"❌ Có lỗi trong quá trình đọc file: {e}")


if __name__ == "__main__":
    main()
